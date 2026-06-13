import openpyxl

# 读取1.xlsx
wb1 = openpyxl.load_workbook('F:\\360MoveData\\Users\\0000\\Desktop\\1.xlsx')
print('=== 1.xlsx ===')
print('工作表:', wb1.sheetnames)
ws1 = wb1.active
print('最大行:', ws1.max_row)
print('最大列:', ws1.max_column)
print('数据预览:')
for row in ws1.iter_rows(min_row=1, max_row=min(10, ws1.max_row), values_only=True):
    print(row)

# 读取1 - 副本.xlsx
wb2 = openpyxl.load_workbook('F:\\360MoveData\\Users\\0000\\Desktop\\1 - 副本.xlsx')
print('\n=== 1 - 副本.xlsx ===')
print('工作表:', wb2.sheetnames)
ws2 = wb2.active
print('最大行:', ws2.max_row)
print('最大列:', ws2.max_column)
print('数据预览:')
for row in ws2.iter_rows(min_row=1, max_row=min(10, ws2.max_row), values_only=True):
    print(row)